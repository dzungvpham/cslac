"""Compute CS-major degree counts per college from IPEDS Completions data.

For each college in ``data/colleges.csv`` this finds the institution's IPEDS
UNITID (via the HD directory file) and sums bachelor's-degree completions in
Computer-Science-related programs for each year 2021-2024.

Which programs count as "CS"
----------------------------
We start from every 6-digit CIP code whose program label contains the word
"Computer" but not "Engineer" (per the user spec), then drop the long tail of
vocational / cross-listed codes that barely appear at liberal-arts colleges by
keeping only codes whose pooled bachelor's completions across *our* college set
clear ``MIN_LAC_TOTAL``. With the 2021-2024 data this yields a clean set:

    11.0701  Computer Science
    11.0101  Computer and Information Sciences, General
    30.0801  Mathematics and Computer Science
    11.0199  Computer and Information Sciences, Other
    11.0899  Computer Software and Media Applications, Other
    11.1003  Computer and Information Systems Security/Information Assurance
    30.3101  Human Computer Interaction
    11.9999  Computer and Information Sciences and Support Services, Other

...and excludes things like Cyber/Computer Forensics, Computer Programming,
Computer Graphics, CNC Machinist, Computer Repair, etc. (the natural break in
the LAC counts falls between ~7 and ~15, so the exact threshold is not
sensitive). The per-college ``cs_cip_codes`` column lists which of these the
college actually awarded, so the choice is auditable.

What gets counted
-----------------
Bachelor's degrees (AWLEVEL = 5), first **and** second major (MAJORNUM 1 + 2),
grand total across demographics (CTOTALT). Including second majors counts every
student who earned a CS major, including the double-majors that are common at
LACs. Both knobs are constants below if a first-major-only count is wanted.

Data sources (IPEDS), cached under ``data/degrees_awarded/raw/``
----------------------------------------------------------------
  - Directory      HD{year}    : UNITID <-> institution name / city / state
  - Completions    C{year}_A   : degrees by 6-digit CIP code
  - Dictionary     C{year}_A   : CIP code -> program label (.xlsx)

Each is fetched from the IPEDS "data-generator" / "dictionary-generator"
endpoint first (recent years only), falling back to the IPEDS data-center bulk
zip (all years). The non-revised ("provisional", HasRV=0) completions file is
used for every year, for consistency -- the most recent year has no revised
release yet. Note the two sources differ in small ways the loaders handle: the
data-center files zero-pad AWLEVEL ("05") and lay the dictionary columns out
differently, so award levels are compared numerically and dictionary columns
are located by header name.

Output: ``data/degrees_awarded.csv`` with columns
    college, UNITID, ipeds_name, cs_cip_codes, <year>, ...
Colleges that could not be matched to a UNITID are still emitted, with blank
UNITID / counts.

Usage:
    python degrees_awarded_scraper.py                  # years 2021-2024
    python degrees_awarded_scraper.py --years 2023 2024
    python degrees_awarded_scraper.py --refresh        # re-download sources
"""

import argparse
import difflib
import io
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
import requests

# --- Paths -----------------------------------------------------------------
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
COLLEGES_CSV = DATA_DIR / "colleges.csv"
RAW_DIR = DATA_DIR / "degrees_awarded" / "raw"
OUTPUT_CSV = DATA_DIR / "degrees_awarded.csv"
CIP_REFERENCE_CSV = DATA_DIR / "degrees_awarded" / "cip_reference.csv"

# --- What to count ---------------------------------------------------------
DEFAULT_YEARS = [2021, 2022, 2023, 2024]
AWLEVEL_BACHELOR = 5          # IPEDS award-level code for a bachelor's degree
MAJORNUMS = (1, 2)            # 1 = first major, 2 = second major
MIN_LAC_TOTAL = 10           # min pooled bachelor's completions (across our
                             # colleges, all years) for a CIP code to count

# --- IPEDS endpoints -------------------------------------------------------
DATA_GENERATOR = "https://nces.ed.gov/ipeds/data-generator"
DICT_GENERATOR = "https://nces.ed.gov/ipeds/dictionary-generator"
DATACENTER = "https://nces.ed.gov/ipeds/datacenter/data"
HTTP_TIMEOUT = 120

# Optional hand overrides for college -> UNITID, if the fuzzy matcher ever
# picks wrong (College Name in colleges.csv -> IPEDS UNITID string).
UNITID_OVERRIDES: dict[str, str] = {}


# ---------------------------------------------------------------------------
# Downloading / extraction
# ---------------------------------------------------------------------------
def _looks_like_zip(content: bytes) -> bool:
    return len(content) > 4 and content[:2] == b"PK"


def _download(url: str, params: dict | None = None) -> bytes | None:
    """GET ``url`` and return the body iff it is a non-empty zip, else None."""
    try:
        resp = requests.get(url, params=params, timeout=HTTP_TIMEOUT)
    except requests.RequestException as exc:
        print(f"    request failed: {exc}")
        return None
    if resp.status_code == 200 and _looks_like_zip(resp.content):
        return resp.content
    return None


def _extract_member(zip_bytes: bytes, suffix: str, exclude_rv: bool = True) -> bytes:
    """Return the bytes of the single ``suffix`` member, skipping *_rv files."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(suffix)]
        if exclude_rv:
            names = [n for n in names if "_rv" not in n.lower()]
        if not names:
            raise ValueError(f"no '{suffix}' member found in zip")
        return zf.read(names[0])


def fetch_source(year: int, kind: str, refresh: bool) -> Path:
    """Download + cache one IPEDS source file, returning the cached path.

    ``kind`` is one of: "hd" (directory csv), "completions" (degrees csv),
    "dictionary" (xlsx). The data-generator endpoint is tried first, then the
    data-center bulk zip as a fallback for years the generator does not serve.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    if kind == "hd":
        out, suffix = RAW_DIR / f"hd{year}.csv", ".csv"
        generator = (DATA_GENERATOR, {"year": year, "tableName": f"HD{year}",
                                      "HasRV": 0, "type": "csv"})
        bulk = f"{DATACENTER}/HD{year}.zip"
    elif kind == "completions":
        out, suffix = RAW_DIR / f"c{year}_a.csv", ".csv"
        generator = (DATA_GENERATOR, {"year": year, "tableName": f"C{year}_A",
                                      "HasRV": 0, "type": "csv"})
        bulk = f"{DATACENTER}/C{year}_A.zip"
    elif kind == "dictionary":
        out, suffix = RAW_DIR / f"c{year}_a.xlsx", ".xlsx"
        generator = (DICT_GENERATOR, {"year": year, "tableName": f"C{year}_A"})
        bulk = f"{DATACENTER}/C{year}_A_Dict.zip"
    else:
        raise ValueError(kind)

    if out.exists() and not refresh:
        return out

    zip_bytes = _download(*generator) or _download(bulk)
    if zip_bytes is None:
        raise RuntimeError(f"could not download {kind} for {year} "
                           f"(tried data-generator and data-center)")
    out.write_bytes(_extract_member(zip_bytes, suffix))
    return out


# ---------------------------------------------------------------------------
# Loading IPEDS tables
# ---------------------------------------------------------------------------
def _read_csv(path: Path, **kwargs) -> pd.DataFrame:
    """Read an IPEDS CSV, tolerating both encodings IPEDS ships.

    Data-generator files are UTF-8 with a BOM; older data-center files (notably
    HD, whose institution names carry accented characters) are Windows-1252.
    utf-8-sig is tried first so the BOM is stripped from column names; on an
    undecodable byte we fall back to cp1252.
    """
    try:
        return pd.read_csv(path, encoding="utf-8-sig", dtype=str, **kwargs)
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="cp1252", dtype=str, **kwargs)


def load_directory(path: Path) -> pd.DataFrame:
    """Load the HD directory: UNITID, INSTNM, IALIAS, CITY, STABBR."""
    return _read_csv(path, usecols=["UNITID", "INSTNM", "IALIAS", "CITY", "STABBR"])


def load_completions(path: Path) -> pd.DataFrame:
    """Load a completions file, normalised for cross-year format drift.

    Older data-center files zero-pad AWLEVEL ("05") and quote fields, so award
    level / major number are coerced to integers and CIP codes are stripped.
    """
    df = _read_csv(path, usecols=["UNITID", "CIPCODE", "MAJORNUM", "AWLEVEL", "CTOTALT"])
    df["CIPCODE"] = df["CIPCODE"].str.strip()
    df["awlevel"] = pd.to_numeric(df["AWLEVEL"], errors="coerce")
    df["majornum"] = pd.to_numeric(df["MAJORNUM"], errors="coerce")
    df["count"] = pd.to_numeric(df["CTOTALT"], errors="coerce").fillna(0).astype(int)
    return df[df["awlevel"] == AWLEVEL_BACHELOR]


def parse_cip_labels(xlsx_path: Path) -> dict[str, str]:
    """Map CIP code -> program label from a completions dictionary .xlsx.

    Handles both dictionary layouts (data-generator vs data-center) by locating
    columns from the header row of the Frequencies sheet rather than by index.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    sheet = next(s for s in wb.sheetnames if s.lower() == "frequencies")
    rows = wb[sheet].iter_rows(values_only=True)
    header = [str(h).lower() if h is not None else "" for h in next(rows)]
    i_var = header.index("varname")
    i_code = header.index("codevalue")
    i_label = header.index("valuelabel")
    labels: dict[str, str] = {}
    for row in rows:
        if row[i_var] and str(row[i_var]).upper() == "CIPCODE" and row[i_code] is not None:
            labels[str(row[i_code]).strip()] = str(row[i_label]).strip()
    return labels


# ---------------------------------------------------------------------------
# College -> UNITID matching
# ---------------------------------------------------------------------------
_NAME_STOPWORDS = {"the"}


def _norm_name(s: str) -> str:
    """Normalise an institution name for matching."""
    if not isinstance(s, str):
        return ""
    s = s.lower().replace("&", " and ")
    s = re.sub(r"\(.*?\)", " ", s)          # drop parentheticals e.g. "(MA)"
    s = re.sub(r"\bsaint\b", "st", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return " ".join(t for t in s.split() if t not in _NAME_STOPWORDS)


def match_unitid(name: str, state: str, city: str, hd_by_state: dict) -> tuple:
    """Best (UNITID, INSTNM, score) for a college within its state, or Nones.

    Tiered scoring within the same state: exact name (100) > exact alias (96) >
    substring either way (88) > token-subset either way (85) > fuzzy ratio. A
    city match adds a small tie-break bonus.
    """
    cn = _norm_name(name)
    ctok = set(cn.split())
    best, best_score = None, -1.0
    for row in hd_by_state.get(state, []):
        hn = row["_norm"]
        htok = set(hn.split())
        aliases = ({_norm_name(a) for a in str(row["IALIAS"]).split("|")}
                   if pd.notna(row["IALIAS"]) else set())
        if hn == cn:
            score = 100.0
        elif cn and cn in aliases:
            score = 96.0
        elif cn and (cn in hn or hn in cn):
            score = 88.0
        elif ctok and (ctok <= htok or htok <= ctok):
            score = 85.0
        else:
            score = difflib.SequenceMatcher(None, cn, hn).ratio() * 80.0
        if str(row["CITY"]).lower() == str(city).lower():
            score += 1.5
        if score > best_score:
            best_score, best = score, row
    if best is not None and best_score >= 70.0:
        return best["UNITID"], best["INSTNM"], round(best_score, 1)
    return None, None, round(best_score, 1)


def match_all_colleges(colleges: pd.DataFrame, hd: pd.DataFrame) -> pd.DataFrame:
    """Attach UNITID / ipeds_name / match score to each college row."""
    hd = hd.copy()
    hd["_norm"] = hd["INSTNM"].map(_norm_name)
    by_state: dict[str, list] = {}
    for row in hd.to_dict("records"):
        by_state.setdefault(row["STABBR"], []).append(row)

    records = []
    for _, c in colleges.iterrows():
        if c["Name"] in UNITID_OVERRIDES:
            uid = UNITID_OVERRIDES[c["Name"]]
            inst = hd.loc[hd["UNITID"] == uid, "INSTNM"]
            records.append({"college": c["Name"], "UNITID": uid,
                            "ipeds_name": inst.iloc[0] if len(inst) else "",
                            "score": 100.0})
            continue
        uid, inst, score = match_unitid(c["Name"], c["State"], c["City"], by_state)
        records.append({"college": c["Name"], "UNITID": uid or "",
                        "ipeds_name": inst or "", "score": score})
    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------
def select_cs_cip_codes(labels: dict[str, str], completions: dict[int, pd.DataFrame],
                        unitids: set[str]) -> tuple[dict[str, str], dict[str, int]]:
    """Choose the CS CIP codes: "Computer" not "Engineer", above MIN_LAC_TOTAL.

    The threshold is applied to bachelor's completions (first + second major)
    pooled across our colleges over every loaded year.
    """
    candidates = {code: lab for code, lab in labels.items()
                  if "computer" in lab.lower() and "engineer" not in lab.lower()
                  and code != "99"}
    pooled: dict[str, int] = {}
    for df in completions.values():
        sub = df[df["UNITID"].isin(unitids) & df["CIPCODE"].isin(candidates)
                 & df["majornum"].isin(MAJORNUMS)]
        for code, total in sub.groupby("CIPCODE")["count"].sum().items():
            pooled[code] = pooled.get(code, 0) + int(total)
    selected = {code: candidates[code] for code in candidates
                if pooled.get(code, 0) >= MIN_LAC_TOTAL}
    # stash pooled totals for the reference file / logging
    selected_pooled = {code: pooled.get(code, 0) for code in selected}
    return dict(sorted(selected.items(), key=lambda kv: -selected_pooled[kv[0]])), selected_pooled


def build_output(colleges_matched: pd.DataFrame, completions: dict[int, pd.DataFrame],
                 cs_codes: dict[str, str], years: list[int]) -> pd.DataFrame:
    """Assemble the per-college, per-year CS-degree table."""
    code_set = set(cs_codes)
    # Per (UNITID, year) -> total CS completions, and per UNITID -> codes used.
    year_counts: dict[int, dict[str, int]] = {y: {} for y in years}
    codes_used: dict[str, set] = {}
    for year, df in completions.items():
        sub = df[df["CIPCODE"].isin(code_set) & df["majornum"].isin(MAJORNUMS)]
        for uid, total in sub.groupby("UNITID")["count"].sum().items():
            year_counts[year][uid] = int(total)
        for uid, code in sub.loc[sub["count"] > 0, ["UNITID", "CIPCODE"]].itertuples(index=False):
            codes_used.setdefault(uid, set()).add(code)

    rows = []
    for _, c in colleges_matched.iterrows():
        uid = c["UNITID"]
        row = {"college": c["college"], "UNITID": uid, "ipeds_name": c["ipeds_name"]}
        if uid:
            present = sorted(codes_used.get(uid, set()))
            row["cs_cip_codes"] = ";".join(present)
            for y in years:
                row[str(y)] = year_counts[y].get(uid, 0)
        else:  # unmatched college: counts unknown
            row["cs_cip_codes"] = ""
            for y in years:
                row[str(y)] = ""
        rows.append(row)
    cols = ["college", "UNITID", "ipeds_name", "cs_cip_codes"] + [str(y) for y in years]
    return pd.DataFrame(rows)[cols]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--years", type=int, nargs="+", default=DEFAULT_YEARS,
                        help="IPEDS data years to include (default: 2021-2024)")
    parser.add_argument("--refresh", action="store_true",
                        help="re-download cached source files")
    args = parser.parse_args()
    years = sorted(args.years)

    colleges = pd.read_csv(COLLEGES_CSV, dtype=str)
    print(f"Colleges to process: {len(colleges)}")

    # 1. Directory (latest requested year is enough -- UNITIDs are stable).
    print(f"\nFetching directory (HD{max(years)})...")
    hd = load_directory(fetch_source(max(years), "hd", args.refresh))

    # 2. Match colleges -> UNITID.
    matched = match_all_colleges(colleges, hd)
    unitids = set(matched.loc[matched["UNITID"] != "", "UNITID"])
    unmatched = matched[matched["UNITID"] == ""]
    print(f"Matched {len(unitids)}/{len(colleges)} colleges to a UNITID.")
    fuzzy = matched[(matched["UNITID"] != "") & (matched["score"] < 90)]
    if not fuzzy.empty:
        print("  Lower-confidence matches (verify):")
        for _, r in fuzzy.iterrows():
            print(f"    [{r['score']:>5}] {r['college']}  ->  {r['ipeds_name']}")
    if not unmatched.empty:
        print(f"  Unmatched ({len(unmatched)}): "
              + ", ".join(unmatched["college"].tolist()))

    # 3. Completions + dictionaries per year.
    completions: dict[int, pd.DataFrame] = {}
    labels: dict[str, str] = {}
    for year in years:
        print(f"\nFetching completions + dictionary for {year}...")
        completions[year] = load_completions(fetch_source(year, "completions", args.refresh))
        labels.update(parse_cip_labels(fetch_source(year, "dictionary", args.refresh)))

    # 4. Pick the CS CIP codes from the "Computer not Engineer" set.
    cs_codes, pooled = select_cs_cip_codes(labels, completions, unitids)
    print(f"\nSelected {len(cs_codes)} CS CIP codes "
          f"(pooled LAC bachelor's >= {MIN_LAC_TOTAL}):")
    for code in cs_codes:
        print(f"    {code}  {pooled[code]:>5}  {cs_codes[code]}")
    CIP_REFERENCE_CSV.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([{"cip_code": c, "label": cs_codes[c], "lac_pooled_total": pooled[c]}
                  for c in cs_codes]).to_csv(CIP_REFERENCE_CSV, index=False)

    # 5. Build + write the per-college table.
    out = build_output(matched, completions, cs_codes, years)
    out.to_csv(OUTPUT_CSV, index=False)
    total = out[[str(y) for y in years]].apply(pd.to_numeric, errors="coerce").sum().sum()
    print(f"\nWrote {OUTPUT_CSV.relative_to(DATA_DIR.parent)} "
          f"({len(out)} colleges, {int(total)} CS degrees {min(years)}-{max(years)}).")
    print(f"Wrote CIP reference -> {CIP_REFERENCE_CSV.relative_to(DATA_DIR.parent)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
