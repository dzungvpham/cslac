"""Maryville College course-schedule scraper.

Maryville publishes each term's section list as a Google-Sheets-hosted
Excel workbook linked from the registrar's "Course Schedules" page. Each
academic year has two workbooks: Fall+Spring and Mayterm+Summer. Sheet
tabs inside each workbook are named after the term (e.g. ``"Fall 2025
Schedule"``, ``"Mayterm 2026 Schedule"``); the term/year are parsed from
the sheet name rather than the link text so the scraper is resilient to
naming variations on the index page.

The workbooks are uploaded ``.xlsx`` files rather than native Google
Sheets, so we fetch them via the
``/export?format=xlsx`` endpoint and parse with ``openpyxl``.

Header conventions differ slightly between the Fall/Spring workbook
(``"CRS CDE" / "CRS TITLE" / "instructor name" / "Meeting Days" / ...``)
and the Mayterm/Summer workbook (``"Instructor"`` capitalized, no
credits, extra ``"SUBTERM CDE"`` column). We map columns by a normalized
header name (lowercased, punctuation stripped) so either layout works.

The course-code cell is a single string like ``"CSC 130 01"`` (subject +
number + section, with internal spacing); we split that into
``course_code`` + ``section``. Maryville lists multi-meeting sections as
separate rows with the same ``CRS CDE`` and different days/times; those
are folded back together per ``(course_code, section)`` and joined with
``;`` in the time column, matching the convention used by the
Self-Service and PowerCampus scrapers.
"""

from __future__ import annotations

import datetime as _dt
import io
import re
import sys
from collections import OrderedDict
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from constants import College  # noqa: E402

from course_schedule.course_schedule_scraper import CourseScheduleScraper

INDEX_URL = "https://www.maryvillecollege.edu/academics/registrar/course-schedules/"

# Only CSC sections matter for our dataset. CRS CDE values look like
# "CSC 130 01" — subject + 1-2 spaces + number + 1-2 spaces + section.
SUBJECTS = ("CSC",)

# Map sheet-name term word -> internal term code. Maryville's "Mayterm"
# is a short intensive term between Spring and Summer; we bucket it as
# "W" (closest analog to a J-term mini-semester).
SHEET_TERM_MAP = [
    ("Mayterm", "W"),
    ("Fall",    "F"),
    ("Spring",  "S"),
    ("Summer",  "Su"),
    ("Winter",  "W"),
]

# Header name -> canonical key. Lower-cased, punctuation stripped before
# lookup so variants like "instructor name" / "Instructor" / "INSTRUCTOR"
# all resolve to the same key.
HEADER_KEYS = {
    "crs cde":       "crs_cde",
    "crs title":     "title",
    "meeting days":  "days",
    "start time":    "start_time",
    "end time":      "end_time",
    "bldg cde":      "bldg",
    "building desc": "bldg_desc",
    "room cde":      "room",
    "instructor":    "instructor",
    "instructor name": "instructor",
}

# Splits "CSC 130 01" / "CSC  130  01" / "ART  140  01" into the subject,
# course number, and section pieces. Subject and section are alphanumeric
# (some Maryville sections look like "01H" or "OL1").
CRS_CDE_RE = re.compile(
    r"^\s*(?P<subject>[A-Z]+)\s+(?P<num>\d+\w*)\s+(?P<section>\w+)\s*$"
)


def _norm_header(s):
    if not s:
        return ""
    return re.sub(r"[^a-z0-9 ]+", "", str(s).lower()).strip()


def _parse_term_year(sheet_name):
    """Map ``"Fall 2025 Schedule"`` -> ``((2025, 2026), "F")``.

    Fall N -> AY (N, N+1); every other term -> AY (N-1, N).
    Returns ``(None, None)`` if no term word + year is found.
    """
    if not sheet_name:
        return None, None
    for word, code in SHEET_TERM_MAP:
        m = re.search(rf"\b{word}\b\s+(\d{{4}})", sheet_name, re.I)
        if m:
            year = int(m.group(1))
            if code == "F":
                ay = (year, year + 1)
            else:
                ay = (year - 1, year)
            return ay, code
    return None, None


def _fmt_time_value(v):
    """Render an openpyxl cell value as ``HH:MM AM/PM``.

    The Fall/Spring sheet stores times as ``datetime.time`` objects;
    Mayterm/Summer cells are often blank or pre-formatted strings.
    """
    if v is None or v == "":
        return ""
    if isinstance(v, _dt.time):
        return v.strftime("%I:%M %p").lstrip("0")
    if isinstance(v, _dt.datetime):
        return v.strftime("%I:%M %p").lstrip("0")
    return str(v).strip()


def _fmt_days(v):
    if v is None:
        return ""
    # The sheet uses single letters separated by spaces ("M W F"); collapse.
    return re.sub(r"\s+", "", str(v).strip())


class MaryvilleScraper(CourseScheduleScraper):
    college = College.MARYVILLE
    fresh_driver_per_load = False
    terms = []
    request_timeout = 60

    def scrape(self):
        session = requests.Session()
        session.headers.update(
            {"User-Agent": "Mozilla/5.0 (cs-lac course-schedule scraper)"}
        )

        resp = session.get(INDEX_URL, timeout=self.request_timeout)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Find unique Google Sheets links — one per academic-year workbook.
        sheet_ids = []
        seen = set()
        for a in soup.find_all("a", href=True):
            href = a["href"]
            m = re.search(r"docs\.google\.com/spreadsheets/d/([\w-]+)", href)
            if not m:
                continue
            sid = m.group(1)
            if sid in seen:
                continue
            seen.add(sid)
            sheet_ids.append((sid, urljoin(INDEX_URL, href)))

        rows = []
        for sid, link_url in sheet_ids:
            export_url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"
            try:
                r = session.get(export_url, timeout=self.request_timeout)
                r.raise_for_status()
            except Exception as e:
                print(f"  failed to fetch workbook {sid}: {e}", flush=True)
                continue
            try:
                wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
            except Exception as e:
                print(f"  failed to parse workbook {sid}: {e}", flush=True)
                continue

            for sheet_name in wb.sheetnames:
                ay, term = _parse_term_year(sheet_name)
                if ay is None or term is None:
                    print(f"  [{sheet_name}] unrecognized sheet name; skipping", flush=True)
                    continue
                label = f"{ay[0]}-{str(ay[1])[-2:]}/{term}"
                page_rows = self._parse_sheet(wb[sheet_name], ay, term, link_url)
                print(f"  [{label}] {len(page_rows)} sections", flush=True)
                rows.extend(page_rows)
        return rows

    def _parse_sheet(self, ws, academic_year, term, url):
        header_idx = None
        headers: dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            mapped = {}
            for col_idx, cell in enumerate(row):
                key = HEADER_KEYS.get(_norm_header(cell))
                if key:
                    mapped[key] = col_idx
            if "crs_cde" in mapped and "title" in mapped:
                header_idx = i
                headers = mapped
                break
        if header_idx is None:
            return []

        # Collapse multi-meeting rows by (course_code, section).
        sections: OrderedDict[tuple, dict] = OrderedDict()
        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            crs_cde = row[headers["crs_cde"]] if headers.get("crs_cde") is not None else None
            m = CRS_CDE_RE.match(str(crs_cde or ""))
            if not m:
                continue
            subject = m.group("subject")
            if subject not in SUBJECTS:
                continue
            course_code = f"{subject} {m.group('num')}"
            section = m.group("section")
            key = (course_code, section)

            def _v(name):
                idx = headers.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            time_text = self._format_meeting(
                _v("days"), _v("start_time"), _v("end_time"), _v("bldg"), _v("room")
            )
            instructor = (_v("instructor") or "")
            instructor = re.sub(r"\s+", " ", str(instructor)).strip()
            title = re.sub(r"\s+", " ", str(_v("title") or "")).strip()

            if key not in sections:
                sections[key] = {
                    "course_code": course_code,
                    "section": section,
                    "course_name": title,
                    "instructor": instructor,
                    "times": [],
                }
            entry = sections[key]
            # Prefer the longer title if the row has one (some rows leave
            # the title blank on continuation lines).
            if title and len(title) > len(entry["course_name"]):
                entry["course_name"] = title
            if instructor and instructor not in entry["instructor"]:
                entry["instructor"] = (
                    f"{entry['instructor']}, {instructor}".strip(", ")
                    if entry["instructor"]
                    else instructor
                )
            if time_text and time_text not in entry["times"]:
                entry["times"].append(time_text)

        rows = []
        for entry in sections.values():
            rows.append(
                self.make_row(
                    academic_year,
                    term,
                    course_code=entry["course_code"],
                    section=entry["section"],
                    course_name=entry["course_name"],
                    instructor=entry["instructor"],
                    time="; ".join(entry["times"]),
                    url=url,
                )
            )
        return rows

    @staticmethod
    def _format_meeting(days, start, end, bldg, room):
        days = _fmt_days(days)
        start = _fmt_time_value(start)
        end = _fmt_time_value(end)
        if start and end:
            time_range = f"{start} - {end}"
        else:
            time_range = start or end
        bits = []
        if days:
            bits.append(days)
        if time_range:
            bits.append(time_range)
        s = " ".join(bits)

        bldg = (str(bldg).strip() if bldg is not None else "")
        room = (str(room).strip() if room is not None else "")
        # Maryville renders unspecified room as "TBA"; building "TBA" is
        # likewise a placeholder rather than a real location.
        if bldg.upper() == "TBA":
            bldg = ""
        if room.upper() == "TBA":
            room = ""
        # Room comes off the sheet as a float (e.g. "201.0"); strip the
        # trailing `.0` so "SSC 201" reads naturally.
        room = re.sub(r"\.0+$", "", room)
        location = " ".join(x for x in [bldg, room] if x)

        if s and location:
            return f"{s} ({location})"
        return s or (f"({location})" if location else "")
